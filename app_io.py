import openpyxl
import copy
import os
import win32com.client
import constants
from docx import *
from utils import *
from slice import DocumentBlock


class ExcelReader:
    """This class helps in reading any given excel sheets."""

    def __init__(self, filename, read_only=True, max_row=None, max_col=None):
        self._workbook = openpyxl.load_workbook(filename, read_only)
        self._sheet = self._workbook.active   # TODO read only the first sheet
        self._max_row = max_row
        self._max_col = max_col

    def fetch_all_rows(self):
        row_end = (self._sheet.max_row if self._max_row is None else self._max_row) + 1
        col_end = (self._sheet.max_column if self._max_col is None else self._max_col) + 1

        # for rx in self._sheet.iter_rows(min_row=2, min_col=1, max_row=row_end, max_col=col_end - 1):
        #     print('AppName: ' + rx[0].value)

        for ri in range(2, row_end):
            row = []
            for ci in range(1, col_end):
                row.append(str(self._sheet.cell(row=ri, column=ci).value).strip())
            yield tuple(row)

    # def get_all_rows(self):
    #     """This method provides the data from excel in the form of a List of Tuples"""
    #     return self._rows


class MasterDocumentHandler:
    """This class handles the operations related to master document"""

    def __init__(self, master_filename):
        self.master_doc = Document(master_filename)

    def create_child_document(self):
        """This method creates and returns 'DocumentBlock' object corresponding to a copy of the master-document."""
        app_doc = copy.deepcopy(self.master_doc)
        doc_block = DocumentBlock(app_doc)
        # return app_doc
        return doc_block

    @staticmethod
    def save(app_name, doc_block):
        output_filename = 'output\\' + format_filename(app_name) + '.docx'
        doc_block.get_doc().save(output_filename)
        word = win32com.client.DispatchEx("Word.Application")
        doc = word.Documents.Open(os.path.abspath(output_filename))

        # canvas = word.ActiveDocument.Shapes(1)
        # for item in canvas.CanvasItems:
        #     print(item) #.TextFrame.TextRange.Text)

        hdr_text = 'ACTIVE DIRECTORY-' + app_name + ' MIGRATION READINESS'
        word.ActiveDocument.Sections(1).Headers(win32com.client.constants.wdHeaderFooterPrimary).Range.Text = hdr_text
        word.ActiveDocument.Save()
        doc.TablesOfContents(1).Update()
        doc.Close(SaveChanges=True)
        word.Quit()
        print('Output file [' + output_filename + '] is created.', end='\n__________________________\n')
